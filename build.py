"""
Regenera index.html a partir de:
  - el .xlsm de circuitos/productividad (descargado de OneDrive/SharePoint)
  - el CSV de control de lotes publicado (Google Sheets)
  - el CSV de pendientes por area publicado (Google Sheets, misma planilla, otra pestaña)

Variables de entorno esperadas (se configuran como Secrets en GitHub Actions):
  XLSM_URL   -> link de descarga directa del .xlsm (con &download=1 al final)
"""
import os
import sys
from datetime import datetime, timezone

import requests
import numpy as np
import pandas as pd
from openpyxl import load_workbook

XLSM_URL = os.environ["XLSM_URL"]
XLSM_LOCAL_PATH = "planilla.xlsm"
TEMPLATE_PATH = "template.html"
OUTPUT_PATH = "index.html"

LOTES_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTx_R6vrK_i25nuoquIzBLw2qEBtL-LQaCeu_VZ7eA1OrbI1HdkqEG-ESOcEPnZLPZeNGFc0F1icxVg/pub?gid=0&single=true&output=csv"
TAREAS_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTx_R6vrK_i25nuoquIzBLw2qEBtL-LQaCeu_VZ7eA1OrbI1HdkqEG-ESOcEPnZLPZeNGFc0F1icxVg/pub?output=csv&gid=442614997"


def descargar_xlsm():
    print("Descargando planilla desde SharePoint...")
    resp = requests.get(XLSM_URL, timeout=60)
    resp.raise_for_status()
    if len(resp.content) < 10_000:
        raise RuntimeError(
            f"La descarga vino demasiado chica ({len(resp.content)} bytes) - "
            "probablemente el link dejo de ser publico o cambio. Revisar XLSM_URL."
        )
    with open(XLSM_LOCAL_PATH, "wb") as f:
        f.write(resp.content)
    print(f"OK: {len(resp.content)} bytes descargados.")


# ---------- 1. AVANCES (V2) + 1. GENERAL -> modulo Construccion de Circuitos ----------

def cargar_avances():
    wb = load_workbook(XLSM_LOCAL_PATH, read_only=True, data_only=True)
    ws = wb["AVANCES (V2)"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c) for c in rows[1]]
    data = [r for r in rows[2:] if r[3] is not None]
    df = pd.DataFrame(data, columns=header)
    return df


def cargar_general_asignaciones():
    """De la hoja '1. GENERAL': columna D (Id Circuito/Alimentador) -> quien esta
    asignado (I, 'Asignado a '), fecha de asignacion (J), tipo de trabajo (K, MT/BT),
    estado (M: En curso / Completado / Publicación parcial / Pendiente) y fecha de
    culminación (N). Un mismo circuito puede tener una fila para MT y otra para BT."""
    wb = load_workbook(XLSM_LOCAL_PATH, read_only=True, data_only=True)
    ws = wb["1. GENERAL"]
    rows = list(ws.iter_rows(values_only=True))
    data = [r for r in rows[2:] if r[3] is not None]

    def fmt_fecha(v):
        if v is None:
            return ""
        try:
            return v.strftime("%d/%m/%Y")
        except AttributeError:
            return str(v)

    asignaciones = {}
    for r in data:
        alimentador = str(r[3]).strip()
        tipo = str(r[10]).strip().upper() if r[10] is not None else ""
        if tipo not in ("MT", "BT"):
            continue
        info = {
            "asignado_a": str(r[8]).strip() if r[8] is not None else "",
            "fecha_asignacion": fmt_fecha(r[9]),
            "fecha_asignacion_raw": r[9] if hasattr(r[9], "strftime") else None,
            "estado": str(r[12]).strip() if r[12] is not None else "",
            "fecha_culminacion": fmt_fecha(r[13]),
            "fecha_culminacion_raw": r[13] if hasattr(r[13], "strftime") else None,
        }
        asignaciones.setdefault(alimentador, {"MT": None, "BT": None})
        asignaciones[alimentador][tipo] = info
    return asignaciones


# ---------- Ritmo de avance: cruce REPORTE_DIARIO x 1. GENERAL ----------

def dias_habiles(fecha_inicio, fecha_fin):
    """Cuenta dias habiles de lunes a viernes (excluye sabado y domingo), inclusive en ambas puntas."""
    if fecha_inicio is None or fecha_fin is None:
        return None
    inicio = np.datetime64(fecha_inicio.date(), "D")
    fin = np.datetime64(fecha_fin.date(), "D") + np.timedelta64(1, "D")
    if fin <= inicio:
        return 0
    return int(np.busday_count(inicio, fin, weekmask="1111100"))


def cargar_reporte_diario_valido():
    wb = load_workbook(XLSM_LOCAL_PATH, read_only=True, data_only=True)
    ws = wb["3.REPORTE_DIARIO"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data = [r for r in rows[1:] if r[0] is not None]
    df = pd.DataFrame(data, columns=header)

    invalidos = {"0", "None", "", "-"}
    df = df[~df["Alimentador / Circuito"].astype(str).str.strip().isin(invalidos)]
    # circuitos con formato de fecha (typo humano tipo "05/07/") no son un circuito real: se excluyen
    df = df[~df["Alimentador / Circuito"].astype(str).str.contains(r"\d{1,2}/\d{1,2}", regex=True, na=False)]
    df = df[df["Tipo de Trabajo"].isin(["MT", "BT"])]

    df["Analista EO"] = df["Analista EO"].astype(str).str.strip().str.capitalize()
    for c in ["Km MT Ajustados", "PD ANDE ", "PD TERCEROS", "PD ANDE Construido", "PD TERCEROS Construido", "Puntos de Servicio"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return df


def calcular_ritmo(df_reporte, asignaciones):
    g = df_reporte.groupby(["Alimentador / Circuito", "Tipo de Trabajo"]).agg(
        analista=("Analista EO", lambda x: ", ".join(sorted(set(x)))),
        km_mt_ajustados=("Km MT Ajustados", "sum"),
        pd_ande=("PD ANDE ", "sum"),
        pd_terceros=("PD TERCEROS", "sum"),
        pd_ande_construido=("PD ANDE Construido", "sum"),
        pd_terceros_construido=("PD TERCEROS Construido", "sum"),
        puntos_servicio=("Puntos de Servicio", "sum"),
    ).reset_index()

    hoy = datetime.now(timezone.utc).replace(tzinfo=None)
    filas = []
    for _, r in g.iterrows():
        circuito = r["Alimentador / Circuito"]
        tipo = r["Tipo de Trabajo"]
        info = (asignaciones.get(circuito) or {}).get(tipo)

        fecha_ini_raw = info["fecha_asignacion_raw"] if info else None
        fecha_fin_raw = (info["fecha_culminacion_raw"] if info and info["estado"] == "Completado" else None) or hoy
        dias = dias_habiles(fecha_ini_raw, fecha_fin_raw) if fecha_ini_raw else None

        def prom(total):
            return round(float(total) / dias, 2) if dias else None

        filas.append({
            "circuito": circuito, "tipo": tipo,
            "analista": (info["asignado_a"] if info else "") or r["analista"],
            "estado": info["estado"] if info else "",
            "fecha_asignacion": info["fecha_asignacion"] if info else "",
            "dias_habiles": dias,
            "km_mt_ajustados": round(float(r["km_mt_ajustados"]), 2),
            "pd_ande": round(float(r["pd_ande"]), 1),
            "pd_terceros": round(float(r["pd_terceros"]), 1),
            "pd_ande_construido": round(float(r["pd_ande_construido"]), 1),
            "pd_terceros_construido": round(float(r["pd_terceros_construido"]), 1),
            "puntos_servicio": round(float(r["puntos_servicio"]), 1),
            "prom_km_mt_ajustados": prom(r["km_mt_ajustados"]),
            "prom_pd_ande": prom(r["pd_ande"]),
            "prom_pd_terceros": prom(r["pd_terceros"]),
            "prom_pd_ande_construido": prom(r["pd_ande_construido"]),
            "prom_pd_terceros_construido": prom(r["pd_terceros_construido"]),
            "prom_puntos_servicio": prom(r["puntos_servicio"]),
        })

    filas.sort(key=lambda f: f["circuito"])
    return filas


def ritmo_data_js(filas):
    def n(v):
        return v if v is not None else "null"

    partes = []
    for f in filas:
        partes.append(
            "{circuito:%r,tipo:%r,analista:%r,estado:%r,fecha_asignacion:%r,dias:%s,"
            "km_mt_ajustados:%s,pd_ande:%s,pd_terceros:%s,pd_ande_construido:%s,"
            "pd_terceros_construido:%s,puntos_servicio:%s,"
            "prom_km_mt_ajustados:%s,prom_pd_ande:%s,prom_pd_terceros:%s,"
            "prom_pd_ande_construido:%s,prom_pd_terceros_construido:%s,prom_puntos_servicio:%s}"
            % (f["circuito"], f["tipo"], f["analista"], f["estado"] or "Sin dato",
               f["fecha_asignacion"] or "&mdash;", n(f["dias_habiles"]),
               f["km_mt_ajustados"], f["pd_ande"], f["pd_terceros"],
               f["pd_ande_construido"], f["pd_terceros_construido"], f["puntos_servicio"],
               n(f["prom_km_mt_ajustados"]), n(f["prom_pd_ande"]), n(f["prom_pd_terceros"]),
               n(f["prom_pd_ande_construido"]), n(f["prom_pd_terceros_construido"]), n(f["prom_puntos_servicio"]))
        )
    return ",\n".join(partes)


def calcular_resumen(df):
    total = len(df)
    en_proceso = (df["Estado General"] == "En proceso").sum()
    no_iniciado = (df["Estado General"] == "No iniciado").sum()
    finalizados = df["Estado General"].astype(str).str.contains("100%").sum()

    regional = (
        df.groupby("Regional")
        .agg(
            circuitos=("Alimentador / Circuito", "count"),
            km_mt_plan=("km MT Plan", "sum"),
            km_mt_real=("km MT Real", "sum"),
            pd_ande_plan=("PD ANDE Plan", "sum"),
            pd_ande_reportado=("PD ANDE Reportado", "sum"),
            pd_tercero_plan=("PD Tercero Plan", "sum"),
        )
        .round(1)
        .sort_values("circuitos", ascending=False)
    )
    return {
        "total": int(total),
        "en_proceso": int(en_proceso),
        "no_iniciado": int(no_iniciado),
        "finalizados": int(finalizados),
        "regional": regional,
    }


def regional_tabla_html(regional_df):
    filas = []
    for reg, r in regional_df.iterrows():
        filas.append(
            f"<tr><td>{reg.title()}</td><td class='num'>{int(r.circuitos)}</td>"
            f"<td class='num'>{r.km_mt_plan:,.0f}</td><td class='num'>{r.km_mt_real:,.0f}</td>"
            f"<td class='num'>{r.pd_ande_plan:,.0f}</td><td class='num'>{r.pd_ande_reportado:,.0f}</td>"
            f"<td class='num'>{r.pd_tercero_plan:,.0f}</td></tr>"
        )
    return "\n".join(filas)


def circuitos_data_js(df, asignaciones):
    cols = ["Regional", "SSEE", "Alimentador / Circuito", "PD ANDE Reportado",
            "PD ANDE Construido BT", "%PD ANDE  BT", "Estado General",
            "Estado Inspeccion Gabinete", "Estado inspeccion 8C",
            "Inicio BT", "Fin BT", "Avance Total MT"]
    d = df[cols].copy()
    d["PD ANDE Reportado"] = pd.to_numeric(d["PD ANDE Reportado"], errors="coerce").fillna(0)
    d["PD ANDE Construido BT"] = pd.to_numeric(d["PD ANDE Construido BT"], errors="coerce").fillna(0)
    d["%PD ANDE  BT"] = pd.to_numeric(d["%PD ANDE  BT"], errors="coerce").fillna(0) * 100
    d["Avance Total MT"] = pd.to_numeric(d["Avance Total MT"], errors="coerce").fillna(0) * 100
    for c in ["Regional", "SSEE", "Alimentador / Circuito", "Estado General",
              "Estado Inspeccion Gabinete", "Estado inspeccion 8C"]:
        d[c] = d[c].fillna("").astype(str).str.strip()

    def fmt_fecha_col(v):
        if pd.isna(v):
            return ""
        try:
            return v.strftime("%d/%m/%Y")
        except AttributeError:
            return str(v)

    filas = []
    for _, r in d.iterrows():
        circuito = r["Alimentador / Circuito"]
        asign = asignaciones.get(circuito.strip(), {"MT": None, "BT": None})
        mt_info = asign.get("MT") or {"asignado_a": "", "fecha_asignacion": "", "estado": "", "fecha_culminacion": ""}
        bt_info = asign.get("BT") or {"asignado_a": "", "fecha_asignacion": "", "estado": "", "fecha_culminacion": ""}
        filas.append(
            "{regional:%r,ssee:%r,circuito:%r,mt:%d,bt:%d,pctbt:%s,estado:%r,"
            "estado_insp:%r,estado_8c:%r,inicio_bt:%r,fin_bt:%r,avance_mt:%s,"
            "mt_asignado:%r,mt_estado:%r,mt_fecha_ini:%r,mt_fecha_fin:%r,"
            "bt_asignado:%r,bt_estado:%r,bt_fecha_ini:%r,bt_fecha_fin:%r}"
            % (r["Regional"] or "Sin asignar", r["SSEE"], circuito,
               int(r["PD ANDE Reportado"]), int(r["PD ANDE Construido BT"]),
               round(float(r["%PD ANDE  BT"]), 2), r["Estado General"] or "No iniciado",
               r["Estado Inspeccion Gabinete"], r["Estado inspeccion 8C"],
               fmt_fecha_col(r["Inicio BT"]), fmt_fecha_col(r["Fin BT"]),
               round(float(r["Avance Total MT"]), 2),
               mt_info["asignado_a"], mt_info["estado"], mt_info["fecha_asignacion"], mt_info["fecha_culminacion"],
               bt_info["asignado_a"], bt_info["estado"], bt_info["fecha_asignacion"], bt_info["fecha_culminacion"])
        )
    return ",\n".join(filas)


# ---------- Control de lotes (Google Sheets publicado) ----------

def cargar_lotes():
    df = pd.read_csv(LOTES_URL, skiprows=3, header=None)
    df = df[df[0].notna()]
    df.columns = list(range(df.shape[1]))

    def col(r, i):
        return str(r[i]).strip() if pd.notna(r[i]) else ""

    filas = []
    for _, r in df.iterrows():
        filas.append({
            "lote": r[2],
            "regional": col(r, 1),
            "estado_general": col(r, 4),          # E
            "postes": int(r[22]) if pd.notna(r[22]) and str(r[22]).strip() not in ("", "-") else None,
            "muestra_ande": col(r, 24),            # Y
            "estado_lote": col(r, 25),             # Z
            "emitido_8a": col(r, 29),              # AD
            "hab8b": col(r, 30),                   # AE
            "emitido_8b": col(r, 31),              # AF
        })

    total = len(filas)

    def z_vacio_o_sin_iniciar(z):
        return z == "" or z.replace("_", " ").strip().lower() == "sin iniciar"

    tramite = sum(1 for f in filas if not z_vacio_o_sin_iniciar(f["estado_lote"]))
    cola = total - tramite
    suma_postes = sum(f["postes"] for f in filas if f["postes"])

    reg_counts = {}
    for f in filas:
        if not z_vacio_o_sin_iniciar(f["estado_lote"]):
            k = f["regional"] or "Sin asignar"
            reg_counts[k] = reg_counts.get(k, 0) + 1

    est_counts = {}
    for f in filas:
        k = f["estado_lote"] or "Sin iniciar"
        est_counts[k] = est_counts.get(k, 0) + 1

    en_curso = [f for f in filas if "curso" in f["estado_lote"].lower()]
    pend_8a = [f for f in filas if f["estado_lote"].lower() == "aceptado" and f["emitido_8a"].upper() != "SI"]
    pend_8b = [f for f in filas if f["hab8b"].upper() == "SI" and f["emitido_8b"].upper() != "SI"]

    return {
        "filas": filas, "total": total, "tramite": tramite, "cola": cola,
        "suma_postes": suma_postes, "reg_counts": reg_counts, "est_counts": est_counts,
        "en_curso": en_curso, "pend_8a": pend_8a, "pend_8b": pend_8b,
    }


def lotes_data_js(lotes):
    filas = []
    for f in lotes["filas"]:
        filas.append(
            "{lote:%r,regional:%r,estado_general:%r,postes:%s,estado_lote:%r,"
            "muestra_ande:%r,emitido_8a:%r,hab8b:%r,emitido_8b:%r}"
            % (f["lote"], f["regional"], f["estado_general"],
               f["postes"] if f["postes"] is not None else "null",
               f["estado_lote"], f["muestra_ande"], f["emitido_8a"], f["hab8b"], f["emitido_8b"])
        )
    return ",\n".join(filas)


def chips_html(filas, vacio_texto="Ninguno por ahora"):
    if not filas:
        return f'<div class="chip-vacio">{vacio_texto}</div>'
    chips = []
    for f in filas:
        reg = f["regional"] or "Sin asignar"
        chips.append(f'<span class="chip">Lote {f["lote"]} <small>({reg})</small></span>')
    return "\n".join(chips)


def bars_html(counts, color_map=None, default_color="#4d9fef"):
    color_map = color_map or {}
    if not counts:
        return ""
    maximo = max(counts.values())
    filas = []
    for k, v in sorted(counts.items(), key=lambda kv: -kv[1]):
        color = color_map.get(k, default_color)
        pct = round(v / maximo * 100) if maximo else 0
        filas.append(
            f'<div class="bar-row"><div class="lbl">{k}</div>'
            f'<div class="bar-track"><div class="bar-fill" style="width:{pct}%; background:{color}"></div></div>'
            f'<div class="bar-val">{v}</div></div>'
        )
    return "\n".join(filas)


# ---------- Pendientes por area (Google Sheets publicado) ----------

def cargar_tareas():
    df = pd.read_csv(TAREAS_URL, header=0)
    df = df[df.iloc[:, 0].notna()]
    df.columns = list(range(df.shape[1]))

    filas = []
    for _, r in df.iterrows():
        filas.append({
            "item": r[0],
            "area": str(r[1]).strip() if pd.notna(r[1]) else "",
            "tarea": str(r[2]).strip() if pd.notna(r[2]) else "",
            "responsable": str(r[4]).strip() if pd.notna(r[4]) else "",
            "regional": str(r[5]).strip() if pd.notna(r[5]) else "",
            "estado": str(r[6]).strip() if pd.notna(r[6]) else "",
            "reporte": str(r[7]).strip() if pd.notna(r[7]) else "",
        })

    def es_ok(estado):
        return estado in ("Realizado", "Completado")

    total = len(filas)
    ok = sum(1 for f in filas if es_ok(f["estado"]))
    pend = total - ok

    por_persona = {}
    for f in filas:
        if not f["responsable"]:
            continue
        por_persona.setdefault(f["responsable"], {"total": 0, "pend": 0})
        por_persona[f["responsable"]]["total"] += 1
        if not es_ok(f["estado"]):
            por_persona[f["responsable"]]["pend"] += 1

    personas_con_pendientes = sum(1 for p in por_persona.values() if p["pend"] > 0)
    areas = sorted({f["area"] for f in filas if f["area"]})

    return {
        "filas": filas, "total": total, "ok": ok, "pend": pend,
        "por_persona": por_persona, "personas_con_pendientes": personas_con_pendientes,
        "areas": areas,
    }


def tareas_data_js(tareas):
    filas = []
    for f in tareas["filas"]:
        filas.append(
            "{item:%r,area:%r,tarea:%r,responsable:%r,regional:%r,estado:%r,reporte:%r}"
            % (f["item"], f["area"], f["tarea"], f["responsable"], f["regional"], f["estado"], f["reporte"])
        )
    return ",\n".join(filas)


def cards_responsables_html(por_persona):
    filas = []
    for nombre, c in sorted(por_persona.items(), key=lambda kv: -kv[1]["pend"]):
        clase = "c-red" if c["pend"] > 0 else "c-green"
        filas.append(
            f'<div class="kpi {clase}"><div class="label">{nombre}</div>'
            f'<div class="value">{c["pend"]}</div>'
            f'<div class="sub">pendientes de {c["total"]} tareas</div></div>'
        )
    return "\n".join(filas)


def filtros_area_html(areas):
    botones = ['<button data-area="Todas" onclick="setFiltroArea(\'Todas\')" class="filtro-btn activo">Todas</button>']
    for a in areas:
        botones.append(f'<button data-area="{a}" onclick="setFiltroArea(\'{a}\')" class="filtro-btn">{a}</button>')
    return "\n".join(botones)


def main():
    descargar_xlsm()
    avances = cargar_avances()
    asignaciones = cargar_general_asignaciones()
    resumen = calcular_resumen(avances)

    print("Descargando control de lotes...")
    lotes = cargar_lotes()
    print(f"OK: {lotes['total']} lotes.")

    print("Descargando pendientes por área...")
    tareas = cargar_tareas()
    print(f"OK: {tareas['total']} tareas.")

    print("Calculando ritmo de avance (REPORTE_DIARIO x 1. GENERAL)...")
    reporte_valido = cargar_reporte_diario_valido()
    ritmo = calcular_ritmo(reporte_valido, asignaciones)
    print(f"OK: {len(ritmo)} combinaciones circuito/tipo con ritmo calculado.")

    with open(TEMPLATE_PATH, encoding="utf-8") as f:
        html = f.read()

    reemplazos = {
        "{{TOTAL_ALIM}}": str(resumen["total"]),
        "{{EN_PROCESO}}": str(resumen["en_proceso"]),
        "{{NO_INICIADO}}": str(resumen["no_iniciado"]),
        "{{FINALIZADOS}}": str(resumen["finalizados"]),
        "{{REGIONAL_TABLE_ROWS}}": regional_tabla_html(resumen["regional"]),
        "{{REGIONAL_CHART_LABELS}}": str([r.title() for r in resumen["regional"].index]),
        "{{REGIONAL_CHART_DATA}}": str(list(resumen["regional"].circuitos.astype(int))),
        "{{CIRCUITOS_DATA_JS}}": circuitos_data_js(avances, asignaciones),
        "{{LOTES_TOTAL}}": str(lotes["total"]),
        "{{LOTES_TRAMITE}}": str(lotes["tramite"]),
        "{{LOTES_COLA}}": str(lotes["cola"]),
        "{{LOTES_POSTES}}": f"{lotes['suma_postes']:,}".replace(",", "."),
        "{{LOTES_BARS_REGIONAL}}": bars_html(lotes["reg_counts"]),
        "{{LOTES_BARS_ESTADO}}": bars_html(lotes["est_counts"], {"Aceptado": "#3ecf8e", "Rechazado": "#f0645f"}, "#f0a94d"),
        "{{LOTES_DATA_JS}}": lotes_data_js(lotes),
        "{{LOTES_EN_CURSO_CHIPS}}": chips_html(lotes["en_curso"], "Ningún lote en curso en la columna Z ahora mismo"),
        "{{LOTES_PEND_8A_CHIPS}}": chips_html(lotes["pend_8a"], "Ningún aceptado esperando firma 8A"),
        "{{LOTES_PEND_8B_CHIPS}}": chips_html(lotes["pend_8b"], "Ningún habilitado esperando firma 8B"),
        "{{TAREAS_TOTAL}}": str(tareas["total"]),
        "{{TAREAS_OK}}": str(tareas["ok"]),
        "{{TAREAS_PEND}}": str(tareas["pend"]),
        "{{TAREAS_PERSONAS}}": str(tareas["personas_con_pendientes"]),
        "{{TAREAS_CARDS_RESPONSABLES}}": cards_responsables_html(tareas["por_persona"]),
        "{{TAREAS_FILTROS_AREA}}": filtros_area_html(tareas["areas"]),
        "{{TAREAS_DATA_JS}}": tareas_data_js(tareas),
        "{{RITMO_DATA_JS}}": ritmo_data_js(ritmo),
        "{{RITMO_TOTAL}}": str(len(ritmo)),
        "{{BUILD_TIMESTAMP}}": datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC"),
    }
    for k, v in reemplazos.items():
        html = html.replace(k, v)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"OK: {OUTPUT_PATH} regenerado con datos frescos.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR en build.py: {e}", file=sys.stderr)
        sys.exit(1)
